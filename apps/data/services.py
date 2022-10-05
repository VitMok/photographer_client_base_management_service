from openpyxl import load_workbook, Workbook
import re

from .models import Client


def _check_phone(phone):
    """ Проверка номера телефона """
    if phone.find('.') != -1:
        phone = phone[:phone.find('.')]
    phone = re.sub("[^0-9]", "", phone)
    if not phone:
        return (False, phone)
    if phone[0] == '9':
        phone = '7' + phone
    else:
        phone = '7' + phone[1:]
    if len(phone) != 11:
        return (False, phone)
    return (True, phone)

def _data_file(file):
    """ Считывание данных из файла Excel
    и занесение в бд """
    wb = load_workbook(file)
    sheets = wb.get_sheet_names()[1:]
    for sheet_name in sheets:
        sheet = wb.get_sheet_by_name(sheet_name)
        for i in range(2, sheet.max_row+1):
            name = str(sheet.cell(row=i, column=2).value)
            phone = str(sheet.cell(row=i, column=3).value)
            correctly, phone = _check_phone(phone)
            if not correctly:
                continue
            date = sheet.cell(row=i, column=4).value
            try:
                total_sum = int(sheet.cell(row=i, column=5).value)
            except:
                total_sum = 0
            email = str(sheet.cell(row=i, column=6).value)
            try:
                client = Client.objects.get(phone=phone)
            except:
                Client.objects.create(
                    name=name,
                    phone=phone,
                    email=email,
                    date=date,
                    total_sum=total_sum
                )
            else:
                client.date = date
                client.total_sum += total_sum
                client.visits_quantity += 1
                client.save()

# def _writing_data_to_excel_file():
#     """  """
#     wb = Workbook()
#     wb.create_sheet(index=1, title="Купили")
#     wb.create_sheet(index=1, title="Нули")
#
#     sheet_bought = wb.get_sheet_by_name("Купили")
#     clients = Client.objects.filter(total_sum__gt=0)
#     row = 2
#     sheet_bought.append(('ИД', 'Имя', 'Телефон', 'Почта', 'Общий чек', 'Всего визитов',
#                          'Дата последнего визита', 'Приглашен на фотосессию'))
#     for client in clients:
#         sent = 'Нет'
#         if client.is_sent:
#             sent = 'Да'
#         sheet_bought.append((row - 1, client.name, client.phone, client.email,
#                              client.total_sum, client.visits_quantity, client.date, sent))
#         row += 1
#     sheet_nulls = wb.get_sheet_by_name("Нули")
#     clients = Client.objects.filter(total_sum=0)
#     row = 2
#     sheet_nulls.append(('ИД', 'Имя', 'Телефон', 'Почта', 'Всего визитов',
#                         'Дата последнего визита', 'Приглашен на фотосессию'))
#     for client in clients:
#         sent = 'Нет'
#         if client.is_sent:
#             sent = 'Да'
#         sheet_nulls.append((row - 1, client.name, client.phone, client.email,
#                             client.visits_quantity, client.date, sent))
#         row += 1
#
#     wb.save(filename="Клиенты фотографов.xlsx")
#     return wb
